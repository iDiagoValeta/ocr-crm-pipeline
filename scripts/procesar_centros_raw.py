"""
Genera y actualiza los catalogos de centros desde los XLSX exportados del CRM.

Entrada principal:
    centros/centrosTablasCRM/{PROVINCIA}.xlsx

Fuentes auxiliares:
    localidades/{PROVINCIA}.json
    localidades/provinceIds.json
    centros/centrosCompletosAntiguos/{PROVINCIA}.txt

Salida:
    centros/{PROVINCIA}.txt
    centros/centrosPendientesMatching.txt
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
CENTROS_DIR = PROJECT_ROOT / "centros"
CRM_DIR = CENTROS_DIR / "centrosTablasCRM"
OLD_CENTROS_DIR = CENTROS_DIR / "centrosCompletosAntiguos"
LOCALIDADES_DIR = PROJECT_ROOT / "localidades"
PROVINCE_IDS_FILE = LOCALIDADES_DIR / "provinceIds.json"
PENDING_FILE = CENTROS_DIR / "centrosPendientesMatching.txt"

DEFAULT_COUNTRY_ID = "95ef97f0-f1d1-e311-9bfe-d89d6765d360"

ID_HEADER_FALLBACK_INDEX = 0
NAME_HEADER = "NOMBRE"
CITY_HEADER = "DIRECCION 1 CIUDAD"


@dataclass
class CenterRecord:
    name: str
    center_id: str
    province_id: str
    city_id: str
    country_id: str


@dataclass
class PendingRecord:
    province: str
    center_id: str
    name: str
    raw_city: str
    reason: str


def normalize_text(text: object) -> str:
    if text is None:
        return ""
    value = str(text).strip().upper()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"[^A-Z0-9 ]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def clean_cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_center_line(line: str) -> Optional[CenterRecord]:
    line = line.strip()
    if not line:
        return None

    try:
        parts = next(csv.reader([line], skipinitialspace=True))
    except csv.Error:
        parts = [p.strip() for p in line.split(",")]

    parts = [p.strip() for p in parts]
    if len(parts) < 5:
        return None

    if len(parts) > 5:
        name = ",".join(parts[:-4]).strip()
        ids = parts[-4:]
    else:
        name = parts[0]
        ids = parts[1:5]

    if not name or not ids[0]:
        return None

    return CenterRecord(
        name=name,
        center_id=ids[0],
        province_id=ids[1],
        city_id=ids[2],
        country_id=ids[3],
    )


def format_center_line(record: CenterRecord) -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="")
    writer.writerow(
        [
            record.name,
            record.center_id,
            record.province_id,
            record.city_id,
            record.country_id,
        ]
    )
    return output.getvalue()


def load_center_records(path: Path) -> List[CenterRecord]:
    records: List[CenterRecord] = []
    if not path.exists():
        return records

    with path.open(encoding="utf-8", errors="replace", newline="") as f:
        for line in f:
            record = parse_center_line(line)
            if record:
                records.append(record)
    return records


def load_center_file(path: Path) -> Dict[str, CenterRecord]:
    records: Dict[str, CenterRecord] = {}
    for record in load_center_records(path):
        records.setdefault(record.center_id.lower(), record)
    return records


def load_province_info() -> Dict[str, Dict[str, str]]:
    if not PROVINCE_IDS_FILE.exists():
        raise FileNotFoundError(f"No existe {PROVINCE_IDS_FILE}")

    with PROVINCE_IDS_FILE.open(encoding="utf-8-sig") as f:
        data = json.load(f)

    available_keys = {
        path.stem.upper()
        for path in LOCALIDADES_DIR.glob("*.json")
        if path.name != "provinceIds.json"
    }
    info: Dict[str, Dict[str, str]] = {}

    for entry in data:
        names = [entry.get("Name", ""), entry.get("Name_cat", "")]
        matching_key = next(
            (
                key
                for key in available_keys
                if normalize_text(key) in {normalize_text(name) for name in names}
            ),
            None,
        )
        if not matching_key:
            continue

        info[matching_key] = {
            "IdProvince": clean_cell(entry.get("IdProvince")),
            "IdCountry": clean_cell(entry.get("CountryId")) or DEFAULT_COUNTRY_ID,
            "Name": clean_cell(entry.get("Name")) or matching_key,
        }

    return info


def city_variants(value: object) -> List[str]:
    raw = clean_cell(value)
    if not raw:
        return []

    variants = {normalize_text(raw)}

    for part in re.split(r"\s*/\s*", raw):
        part_norm = normalize_text(part)
        if part_norm:
            variants.add(part_norm)

    article_suffix = re.match(r"(.+)\s+\((EL|LA|LOS|LAS)\)$", raw, flags=re.I)
    if article_suffix:
        name, article = article_suffix.groups()
        variants.add(normalize_text(f"{article} {name}"))
        variants.add(normalize_text(name))

    article_comma = re.match(r"(.+),\s*(EL|LA|LOS|LAS)$", raw, flags=re.I)
    if article_comma:
        name, article = article_comma.groups()
        variants.add(normalize_text(f"{article} {name}"))
        variants.add(normalize_text(name))

    return [v for v in variants if v]


CITY_ALIASES = {
    "ALICANTE ALACANT": "ALICANTE",
    "ALACANT": "ALICANTE",
    "ELCHE": "ELCHE ELX",
    "ELX": "ELCHE ELX",
    "SAN VICENTE DEL RASPEIG SANT VICENT DEL RASPEIG": "SAN VICENTE DEL RASPEIG",
    "CASTELLON": "CASTELLON DE LA PLANA",
    "CASTELLO": "CASTELLON DE LA PLANA",
    "CASTELLON CASTELLO": "CASTELLON DE LA PLANA",
    "VILLARREAL": "VILA REAL",
    "VILLA REAL": "VILA REAL",
    "ALMAZORA": "ALMASSORA",
    "ALQUERIAS DEL NINO PERDIDO": "ALQUERIES DEL NINO PERDIDO",
    "SAGUNTO": "SAGUNTO SAGUNT",
    "SAGUNT": "SAGUNTO SAGUNT",
    "JATIVA": "XATIVA",
    "ALCIRA": "ALZIRA",
    "ALCOY": "ALCOI",
    "PALMA DE MALLORCA": "PALMA",
}


def expand_city_lookup_keys(value: object) -> List[str]:
    keys = city_variants(value)
    expanded = list(keys)
    for key in keys:
        alias = CITY_ALIASES.get(key)
        if alias:
            expanded.extend(city_variants(alias))
    return list(dict.fromkeys(expanded))


def load_city_maps(province_key: str) -> Tuple[Dict[str, dict], Dict[str, dict]]:
    path = LOCALIDADES_DIR / f"{province_key}.json"
    if not path.exists():
        return {}, {}

    with path.open(encoding="utf-8-sig") as f:
        cities = json.load(f)

    by_name: Dict[str, dict] = {}
    by_id: Dict[str, dict] = {}

    for city in cities:
        city_id = clean_cell(city.get("IdCity"))
        if city_id:
            by_id[city_id.lower()] = city

        for key_name in (city.get("Name"), city.get("Name_cat")):
            for variant in expand_city_lookup_keys(key_name):
                by_name.setdefault(variant, city)

    return by_name, by_id


def find_city(raw_city: str, city_by_name: Dict[str, dict]) -> Optional[dict]:
    for key in expand_city_lookup_keys(raw_city):
        city = city_by_name.get(key)
        if city:
            return city
    return None


def display_city_name(city: Optional[dict], fallback: str) -> str:
    if city:
        return clean_cell(city.get("Name")) or clean_cell(city.get("Name_cat")) or fallback
    return fallback


def has_trailing_location(name: str, city_name: str, province_key: str) -> bool:
    matches = re.findall(r"\(([^)]+)\)\s*$", name)
    if not matches:
        return False
    last = normalize_text(matches[-1])
    city_norm = normalize_text(city_name)
    province_norm = normalize_text(province_key)
    return bool(
        last == city_norm
        or city_norm in last
        or last in city_norm
        or last == province_norm
    )


def name_with_location(name: str, city_name: str, province_key: str) -> str:
    cleaned = re.sub(r"\s+", " ", name).strip()
    location = clean_cell(city_name) or province_key
    if has_trailing_location(cleaned, location, province_key):
        return cleaned
    return f"{cleaned} ({location.upper()})"


def row_header_map(headers: Sequence[object]) -> Dict[str, int]:
    return {normalize_text(header): index for index, header in enumerate(headers)}


def read_crm_rows(path: Path) -> Iterable[Tuple[str, str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        return

    header_map = row_header_map(headers)
    name_index = header_map.get(NAME_HEADER)
    city_index = header_map.get(CITY_HEADER)

    if name_index is None:
        raise ValueError(f"{path.name}: no se encuentra la columna 'Nombre'")

    for row in rows:
        center_id = clean_cell(row[ID_HEADER_FALLBACK_INDEX] if len(row) > ID_HEADER_FALLBACK_INDEX else "")
        name = clean_cell(row[name_index] if len(row) > name_index else "")
        raw_city = clean_cell(row[city_index] if city_index is not None and len(row) > city_index else "")
        if center_id or name:
            yield center_id, name, raw_city


def build_record_from_raw(
    province_key: str,
    center_id: str,
    name: str,
    raw_city: str,
    province_info: Dict[str, str],
    city_by_name: Dict[str, dict],
    city_by_id: Dict[str, dict],
    old_records: Dict[str, CenterRecord],
) -> Tuple[Optional[CenterRecord], Optional[str]]:
    old = old_records.get(center_id.lower())
    city = find_city(raw_city, city_by_name) if raw_city else None

    if not city and old and old.city_id:
        city = city_by_id.get(old.city_id.lower())
        if not city:
            return (
                CenterRecord(
                    name=name_with_location(name, province_key, province_key),
                    center_id=center_id,
                    province_id=province_info["IdProvince"],
                    city_id=old.city_id,
                    country_id=province_info.get("IdCountry") or old.country_id or DEFAULT_COUNTRY_ID,
                ),
                None,
            )

    if not city:
        reason = "sin ciudad en RAW ni respaldo antiguo" if not raw_city else "ciudad RAW no encontrada en localidades"
        return None, reason

    city_name = display_city_name(city, raw_city or province_key)
    return (
        CenterRecord(
            name=name_with_location(name, city_name, province_key),
            center_id=center_id,
            province_id=clean_cell(city.get("IdProvince")) or province_info["IdProvince"],
            city_id=clean_cell(city.get("IdCity")),
            country_id=province_info.get("IdCountry") or DEFAULT_COUNTRY_ID,
        ),
        None,
    )


def sort_records(records: Iterable[CenterRecord]) -> List[CenterRecord]:
    return sorted(records, key=lambda record: (normalize_text(record.name), record.center_id.lower()))


def write_center_file(path: Path, records: Iterable[CenterRecord]) -> None:
    lines = [format_center_line(record) for record in sort_records(records)]
    with path.open("w", encoding="utf-8", newline="") as f:
        f.write("\n".join(lines))
        if lines:
            f.write("\n")


def write_pending_file(path: Path, pending: Sequence[PendingRecord]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Province", "Id", "Name", "RawCity", "Reason"])
        for item in sorted(pending, key=lambda p: (p.province, normalize_text(p.name), p.center_id)):
            writer.writerow([item.province, item.center_id, item.name, item.raw_city, item.reason])


def process_province(path: Path, province_info: Dict[str, Dict[str, str]]) -> Tuple[str, List[CenterRecord], List[PendingRecord], Dict[str, int]]:
    province_key = path.stem.upper()
    if province_key not in province_info:
        raise ValueError(f"{province_key}: no existe IdProvince en {PROVINCE_IDS_FILE}")

    current_records = load_center_records(CENTROS_DIR / f"{province_key}.txt")
    current_ids = {record.center_id.lower() for record in current_records}
    old_records = load_center_file(OLD_CENTROS_DIR / f"{province_key}.txt")
    city_by_name, city_by_id = load_city_maps(province_key)

    output_records = list(current_records)
    pending: List[PendingRecord] = []
    stats = {
        "raw": 0,
        "preserved": 0,
        "added": 0,
        "pending": 0,
        "missing_required": 0,
    }

    for center_id, name, raw_city in read_crm_rows(path):
        stats["raw"] += 1
        if not center_id or not name:
            stats["missing_required"] += 1
            pending.append(
                PendingRecord(province_key, center_id or "(sin id)", name or "(sin nombre)", raw_city, "faltan Id o Nombre")
            )
            continue

        key = center_id.lower()
        if key in current_ids:
            stats["preserved"] += 1
            continue

        record, reason = build_record_from_raw(
            province_key=province_key,
            center_id=center_id,
            name=name,
            raw_city=raw_city,
            province_info=province_info[province_key],
            city_by_name=city_by_name,
            city_by_id=city_by_id,
            old_records=old_records,
        )
        if record:
            output_records.append(record)
            current_ids.add(key)
            stats["added"] += 1
        else:
            pending.append(PendingRecord(province_key, center_id, name, raw_city, reason or "no resuelto"))
            stats["pending"] += 1

    return province_key, output_records, pending, stats


def selected_xlsx_files(selected: Sequence[str]) -> List[Path]:
    if not CRM_DIR.exists():
        raise FileNotFoundError(f"No existe {CRM_DIR}")

    selected_keys = {province.upper() for province in selected}
    files = sorted(CRM_DIR.glob("*.xlsx"))
    if selected_keys:
        files = [path for path in files if path.stem.upper() in selected_keys]
    return files


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Procesa centros CRM por provincia.")
    parser.add_argument("--dry-run", action="store_true", help="Calcula resultados sin escribir archivos.")
    parser.add_argument("--province", action="append", default=[], help="Procesa solo esta provincia. Repetible.")
    args = parser.parse_args(argv)

    province_info = load_province_info()
    files = selected_xlsx_files(args.province)
    if not files:
        print("No hay XLSX que procesar.")
        return 1

    all_pending: List[PendingRecord] = []
    processed: List[Tuple[str, List[CenterRecord], Dict[str, int]]] = []

    for path in files:
        province_key, records, pending, stats = process_province(path, province_info)
        processed.append((province_key, records, stats))
        all_pending.extend(pending)

    if not args.dry_run:
        CENTROS_DIR.mkdir(parents=True, exist_ok=True)
        for province_key, records, _ in processed:
            write_center_file(CENTROS_DIR / f"{province_key}.txt", records)
        write_pending_file(PENDING_FILE, all_pending)

    print("=" * 80)
    print("RESULTADOS" + (" (dry-run)" if args.dry_run else ""))
    print("=" * 80)
    for province_key, records, stats in processed:
        print(
            f"{province_key}.txt -> total={len(records)} "
            f"raw={stats['raw']} preserved={stats['preserved']} "
            f"added={stats['added']} pending={stats['pending']}"
        )
    print(f"Pendientes matching -> {len(all_pending)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
